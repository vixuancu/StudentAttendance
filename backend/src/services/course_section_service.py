from datetime import date, datetime, time, timedelta
from io import BytesIO
import re
import unicodedata
from typing import Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy.exc import IntegrityError

from src.constant.error_code import ERROR_CODES
from src.constant.schedule_period import PERIOD_1_START_TIME, PERIOD_DURATION_MINUTES
from src.db.models.enums import SessionStatus
from src.db.models.course_section import CourseSection
from src.db.models.student import Student
from src.dto.common import PaginationParams
from src.dto.request.course_section_request import (
    CourseSectionEnrollmentCreateRequest,
    CourseSectionCreateRequest,
    CourseSectionUpdateRequest,
)
from src.dto.response.student_response import (
    StudentImportErrorResponse,
    StudentImportResultResponse,
)
from src.repository.interfaces.i_course_section_repo import ICourseSectionRepository
from src.services.interfaces.i_course_section_service import ICourseSectionService
from src.utils.exception import AlreadyExists, NotFound, Validation


class CourseSectionService(ICourseSectionService):

    HEADER_ALIASES = {
        "student_code": {
            "masinhvien",
            "mssv",
            "studentcode",
            "code",
        }
    }

    def __init__(self, repo: ICourseSectionRepository):
        self.repo = repo

    @staticmethod
    def _is_name_unique_violation(exc: IntegrityError) -> bool:
        message = str(getattr(exc, "orig", exc)).lower()
        return "course_section_name_key" in message or (
            "duplicate key value violates unique constraint" in message
            and "(name)" in message
        )

    @staticmethod
    def _normalize_text(value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value.strip().lower())
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r"[^a-z0-9]+", "", normalized)
        return normalized

    @staticmethod
    def _stringify_cell(value) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _build_header_index(self, headers: list[str]) -> dict[str, int]:
        mapped: dict[str, int] = {}
        for idx, raw in enumerate(headers):
            key = self._normalize_text(raw)
            if not key:
                continue
            for column, aliases in self.HEADER_ALIASES.items():
                if key in aliases and column not in mapped:
                    mapped[column] = idx
                    break
        return mapped

    @staticmethod
    def _validate_file_extension(filename: str | None) -> None:
        if not filename:
            return
        lower_name = filename.lower()
        if lower_name.endswith(".xlsx"):
            return
        if lower_name.endswith(".xls"):
            raise Validation(
                message="Định dạng .xls chưa được hỗ trợ, vui lòng lưu file dưới dạng .xlsx"
            )
        raise Validation(message="Chỉ hỗ trợ file Excel định dạng .xlsx")

    @staticmethod
    def _append_import_error(
        errors: list[StudentImportErrorResponse],
        row: int,
        field: str,
        message: str,
        student_code: str | None = None,
    ) -> None:
        errors.append(
            StudentImportErrorResponse(
                row=row,
                field=field,
                student_code=student_code,
                message=message,
            )
        )

    async def list_sections(
        self,
        pagination: PaginationParams,
        search: Optional[str] = None,
        is_cancel: Optional[bool] = None,
        lecturer_id: Optional[int] = None,
    ) -> tuple[list[tuple[CourseSection, int]], int]:
        items = await self.repo.list_sections(
            skip=pagination.offset,
            limit=pagination.limit,
            search=search,
            is_cancel=is_cancel,
            lecturer_id=lecturer_id,
        )
        total = await self.repo.count_sections(
            search=search,
            is_cancel=is_cancel,
            lecturer_id=lecturer_id,
        )
        return items, total

    async def get_by_id(self, section_id: int) -> CourseSection:
        item = await self.repo.get_by_id(section_id)
        if item is None:
            raise NotFound(ERROR_CODES.COURSE_SECTION.COURSE_SECTION_NOT_FOUND)
        return item

    async def _validate_foreign_keys(
        self,
        course_id: Optional[int],
        user_id: Optional[int],
        room_id: Optional[int],
    ):
        if course_id is not None:
            course = await self.repo.get_course_by_id(course_id)
            if course is None:
                raise NotFound(ERROR_CODES.COURSE.COURSE_NOT_FOUND)

        if user_id is not None:
            lecturer = await self.repo.get_lecturer_by_id(user_id)
            if lecturer is None:
                raise NotFound(ERROR_CODES.COURSE_SECTION.LECTURER_NOT_FOUND)

        if room_id is not None:
            room = await self.repo.get_room_by_id(room_id)
            if room is None:
                raise NotFound(ERROR_CODES.CLASSROOM.CLASSROOM_NOT_FOUND)

    async def _validate_schedule_users(
        self, schedules: list[dict], default_user_id: int
    ) -> None:
        user_ids = {int(item.get("user_id") or default_user_id) for item in schedules}
        for user_id in user_ids:
            lecturer = await self.repo.get_lecturer_by_id(user_id)
            if lecturer is None:
                raise NotFound(ERROR_CODES.COURSE_SECTION.LECTURER_NOT_FOUND)

    @staticmethod
    def _validate_time_fields(
        start_date: datetime,
        end_date: datetime,
        start_time: Optional[datetime],
        end_time: Optional[datetime],
    ):
        if start_date >= end_date:
            raise Validation(
                message="Ngày bắt đầu phải nhỏ hơn ngày kết thúc",
                error_code=ERROR_CODES.COURSE_SECTION.INVALID_DATE_RANGE,
            )

        if start_time is not None and end_time is not None and start_time >= end_time:
            raise Validation(
                message="Giờ bắt đầu phải nhỏ hơn giờ kết thúc",
                error_code=ERROR_CODES.COURSE_SECTION.INVALID_TIME_RANGE,
            )

    @staticmethod
    def _compute_period_times(
        *,
        reference_date: datetime,
        start_period: int,
        number_of_periods: int,
    ) -> tuple[datetime, datetime]:
        if start_period < 1:
            raise Validation(message="Tiết bắt đầu phải lớn hơn hoặc bằng 1")
        if number_of_periods < 1:
            raise Validation(message="Số tiết phải lớn hơn hoặc bằng 1")

        period_1_start = datetime.combine(reference_date.date(), PERIOD_1_START_TIME)
        start_time = period_1_start + timedelta(
            minutes=(start_period - 1) * PERIOD_DURATION_MINUTES
        )
        end_time = start_time + timedelta(
            minutes=number_of_periods * PERIOD_DURATION_MINUTES
        )
        return start_time, end_time

    def _apply_period_times(
        self,
        *,
        reference_date: datetime,
        schedules: list[dict],
    ) -> list[dict]:
        normalized: list[dict] = []
        for item in schedules:
            start_time, end_time = self._compute_period_times(
                reference_date=reference_date,
                start_period=int(item["start_period"]),
                number_of_periods=int(item["number_of_periods"]),
            )
            payload = dict(item)
            payload["start_time"] = start_time
            payload["end_time"] = end_time
            normalized.append(payload)
        return normalized

    @staticmethod
    def _to_python_weekday(day_of_week: int) -> int:
        if day_of_week == 8:
            return 6
        return day_of_week - 2

    def _build_session_payloads(
        self,
        *,
        section_id: int,
        start_date: datetime,
        end_date: datetime,
        schedules: list[dict],
    ) -> list[dict]:
        start_day = start_date.date()
        end_day = end_date.date()
        generated: list[dict] = []

        for schedule in schedules:
            target_weekday = self._to_python_weekday(int(schedule["day_of_week"]))
            cursor = start_day
            while cursor <= end_day:
                if cursor.weekday() == target_weekday:
                    start_time, end_time = self._compute_period_times(
                        reference_date=datetime.combine(cursor, time.min),
                        start_period=int(schedule["start_period"]),
                        number_of_periods=int(schedule["number_of_periods"]),
                    )
                    generated.append(
                        {
                            "course_section_id": section_id,
                            "room_id": schedule.get("room_id"),
                            "session_date": datetime.combine(cursor, time.min),
                            "start_time": start_time,
                            "end_time": end_time,
                            "late_time": None,
                            "status": SessionStatus.PENDING,
                            "note": None,
                            "is_cancel": False,
                        }
                    )
                cursor += timedelta(days=1)

        return generated

    @staticmethod
    def _session_key(
        payload: dict,
    ) -> tuple[date, Optional[int], Optional[datetime], Optional[datetime]]:
        session_date = payload["session_date"].date()
        room_id = payload.get("room_id")
        start_time = payload.get("start_time")
        end_time = payload.get("end_time")
        return (session_date, room_id, start_time, end_time)

    async def _sync_class_sessions(
        self,
        *,
        section_id: int,
        start_date: datetime,
        end_date: datetime,
        schedules: list[dict],
    ) -> None:
        target_sessions = self._build_session_payloads(
            section_id=section_id,
            start_date=start_date,
            end_date=end_date,
            schedules=schedules,
        )
        if not target_sessions:
            return

        existing = await self.repo.list_class_sessions(section_id)
        existing_keys = {
            (
                item.session_date.date(),
                item.room_id,
                item.start_time,
                item.end_time,
            )
            for item in existing
        }

        to_create = [
            payload
            for payload in target_sessions
            if self._session_key(payload) not in existing_keys
        ]

        await self.repo.create_class_sessions(to_create)

    @staticmethod
    def _weekday_label(day_of_week: int) -> str:
        labels = {
            2: "Thứ 2",
            3: "Thứ 3",
            4: "Thứ 4",
            5: "Thứ 5",
            6: "Thứ 6",
            7: "Thứ 7",
            8: "Chủ nhật",
        }
        return labels.get(day_of_week, f"Thứ {day_of_week}")

    @staticmethod
    def _format_period_range(
        start_period: int, number_of_periods: int
    ) -> dict[str, int]:
        end_period = start_period + number_of_periods - 1
        return {
            "start": start_period,
            "end": end_period,
        }

    def _build_conflict_details(
        self,
        *,
        conflict_type: str,
        request_user_id: int,
        request_room_id: int,
        request_day_of_week: int,
        request_start_date: datetime,
        request_end_date: datetime,
        request_start_period: int,
        request_number_of_periods: int,
        conflict_section: CourseSection,
    ) -> dict:
        requested_period = self._format_period_range(
            request_start_period,
            request_number_of_periods,
        )
        conflict_period = self._format_period_range(
            conflict_section.start_period,
            conflict_section.number_of_periods,
        )
        return {
            "conflict_type": conflict_type,
            "requested": {
                "user_id": request_user_id,
                "room_id": request_room_id,
                "day_of_week": request_day_of_week,
                "day_of_week_label": self._weekday_label(request_day_of_week),
                "date_range": {
                    "start": request_start_date.date().isoformat(),
                    "end": request_end_date.date().isoformat(),
                },
                "period_range": requested_period,
                "period_text": f"Tiết {requested_period['start']}-{requested_period['end']}",
            },
            "conflict": {
                "id": conflict_section.id,
                "name": conflict_section.name,
                "user_id": conflict_section.user_id,
                "room_id": conflict_section.room_id,
                "day_of_week": conflict_section.day_of_week,
                "day_of_week_label": self._weekday_label(conflict_section.day_of_week),
                "date_range": {
                    "start": conflict_section.start_date.date().isoformat(),
                    "end": conflict_section.end_date.date().isoformat(),
                },
                "period_range": conflict_period,
                "period_text": f"Tiết {conflict_period['start']}-{conflict_period['end']}",
            },
        }

    async def _validate_schedule_conflicts(
        self,
        *,
        user_id: int,
        room_id: int,
        day_of_week: int,
        start_date: datetime,
        end_date: datetime,
        start_period: int,
        number_of_periods: int,
        exclude_section_id: Optional[int] = None,
    ) -> None:
        lecturer_conflict = await self.repo.get_lecturer_schedule_conflict(
            user_id=user_id,
            day_of_week=day_of_week,
            start_date=start_date,
            end_date=end_date,
            start_period=start_period,
            number_of_periods=number_of_periods,
            exclude_section_id=exclude_section_id,
        )
        if lecturer_conflict is not None:
            raise Validation(
                message=(
                    "Giảng viên đã có lớp bị trùng lịch "
                    f"(lớp: {lecturer_conflict.name}, mã: {lecturer_conflict.id})"
                ),
                error_code=ERROR_CODES.COURSE_SECTION.LECTURER_SCHEDULE_CONFLICT,
                details=self._build_conflict_details(
                    conflict_type="lecturer",
                    request_user_id=user_id,
                    request_room_id=room_id,
                    request_day_of_week=day_of_week,
                    request_start_date=start_date,
                    request_end_date=end_date,
                    request_start_period=start_period,
                    request_number_of_periods=number_of_periods,
                    conflict_section=lecturer_conflict,
                ),
            )

    @staticmethod
    def _is_period_overlap(
        start_period_a: int,
        number_of_periods_a: int,
        start_period_b: int,
        number_of_periods_b: int,
    ) -> bool:
        end_a = start_period_a + number_of_periods_a
        end_b = start_period_b + number_of_periods_b
        return start_period_a < end_b and end_a > start_period_b

    def _validate_internal_schedule_items(self, schedules: list[dict]) -> None:
        for i in range(len(schedules)):
            left = schedules[i]
            for j in range(i + 1, len(schedules)):
                right = schedules[j]
                if left["day_of_week"] != right["day_of_week"]:
                    continue
                if not self._is_period_overlap(
                    left["start_period"],
                    left["number_of_periods"],
                    right["start_period"],
                    right["number_of_periods"],
                ):
                    continue
                if left["room_id"] == right["room_id"]:
                    raise Validation(
                        message="Danh sách lịch học trong cùng lớp bị trùng phòng và trùng tiết",
                    )
                raise Validation(
                    message="Danh sách lịch học trong cùng lớp bị trùng tiết",
                )

    def _normalize_schedules_from_create(
        self,
        request: CourseSectionCreateRequest,
    ) -> list[dict]:
        if request.schedules:
            return [
                {
                    "day_of_week": item.day_of_week,
                    "start_period": item.start_period,
                    "number_of_periods": item.number_of_periods,
                    "start_time": item.start_time,
                    "end_time": item.end_time,
                    "room_id": item.room_id or request.room_id,
                    "user_id": item.user_id or request.user_id,
                }
                for item in request.schedules
            ]

        return [
            {
                "day_of_week": request.day_of_week,
                "start_period": request.start_period,
                "number_of_periods": request.number_of_periods,
                "start_time": request.start_time,
                "end_time": request.end_time,
                "room_id": request.room_id,
                "user_id": request.user_id,
            }
        ]

    def _normalize_schedules_from_update(
        self,
        *,
        request: CourseSectionUpdateRequest,
        item: CourseSection,
        data: dict,
    ) -> list[dict]:
        if request.schedules:
            default_room_id = data.get("room_id", item.room_id)
            default_user_id = data.get("user_id", item.user_id)
            return [
                {
                    "day_of_week": s.day_of_week,
                    "start_period": s.start_period,
                    "number_of_periods": s.number_of_periods,
                    "start_time": s.start_time,
                    "end_time": s.end_time,
                    "room_id": s.room_id or default_room_id,
                    "user_id": s.user_id or default_user_id,
                }
                for s in request.schedules
            ]

        legacy_fields = {
            "day_of_week",
            "start_period",
            "number_of_periods",
            "start_time",
            "end_time",
            "room_id",
        }
        if any(field in data for field in legacy_fields):
            return [
                {
                    "day_of_week": data.get("day_of_week", item.day_of_week),
                    "start_period": data.get("start_period", item.start_period),
                    "number_of_periods": data.get(
                        "number_of_periods", item.number_of_periods
                    ),
                    "start_time": data.get("start_time", item.start_time),
                    "end_time": data.get("end_time", item.end_time),
                    "room_id": data.get("room_id", item.room_id),
                    "user_id": data.get("user_id", item.user_id),
                }
            ]

        existing = list(item.schedules or [])
        if existing:
            return [
                {
                    "day_of_week": s.day_of_week,
                    "start_period": s.start_period,
                    "number_of_periods": s.number_of_periods,
                    "start_time": s.start_time,
                    "end_time": s.end_time,
                    "room_id": s.room_id,
                    "user_id": s.user_id,
                }
                for s in existing
                if not s.is_cancel
            ]

        return [
            {
                "day_of_week": item.day_of_week,
                "start_period": item.start_period,
                "number_of_periods": item.number_of_periods,
                "start_time": item.start_time,
                "end_time": item.end_time,
                "room_id": item.room_id,
                "user_id": item.user_id,
            }
        ]

    async def _validate_schedule_rooms(
        self,
        *,
        schedules: list[dict],
        start_date: datetime,
        end_date: datetime,
        default_user_id: int,
        exclude_section_id: Optional[int] = None,
    ) -> None:
        room_ids = {
            int(item["room_id"])
            for item in schedules
            if item.get("room_id") is not None
        }
        for room_id in room_ids:
            room = await self.repo.get_room_by_id(room_id)
            if room is None:
                raise NotFound(ERROR_CODES.CLASSROOM.CLASSROOM_NOT_FOUND)

        for schedule in schedules:
            room_id = int(schedule["room_id"])
            day_of_week = int(schedule["day_of_week"])
            start_period = int(schedule["start_period"])
            number_of_periods = int(schedule["number_of_periods"])
            request_user_id = int(schedule.get("user_id") or default_user_id)

            room_conflict = await self.repo.get_room_schedule_conflict(
                room_id=room_id,
                day_of_week=day_of_week,
                start_date=start_date,
                end_date=end_date,
                start_period=start_period,
                number_of_periods=number_of_periods,
                exclude_section_id=exclude_section_id,
            )
            if room_conflict is not None:
                raise Validation(
                    message=(
                        "Phòng học đã có lớp bị trùng lịch "
                        f"(lớp: {room_conflict.name}, mã: {room_conflict.id})"
                    ),
                    error_code=ERROR_CODES.COURSE_SECTION.ROOM_SCHEDULE_CONFLICT,
                    details=self._build_conflict_details(
                        conflict_type="room",
                        request_user_id=request_user_id,
                        request_room_id=room_id,
                        request_day_of_week=day_of_week,
                        request_start_date=start_date,
                        request_end_date=end_date,
                        request_start_period=start_period,
                        request_number_of_periods=number_of_periods,
                        conflict_section=room_conflict,
                    ),
                )

    async def create(self, request: CourseSectionCreateRequest) -> CourseSection:
        normalized_name = request.name.strip()
        existed = await self.repo.get_by_name_ci(normalized_name)
        if existed is not None:
            raise AlreadyExists(
                ERROR_CODES.COURSE_SECTION.COURSE_SECTION_NAME_IS_EXISTED
            )

        await self._validate_foreign_keys(
            course_id=request.course_id,
            user_id=request.user_id,
            room_id=request.room_id,
        )

        schedules = self._normalize_schedules_from_create(request)
        if not schedules:
            raise Validation(message="Lớp tín chỉ phải có ít nhất một lịch học")
        schedules = self._apply_period_times(
            reference_date=request.start_date,
            schedules=schedules,
        )

        self._validate_internal_schedule_items(schedules)
        await self._validate_schedule_users(schedules, request.user_id)
        await self._validate_schedule_rooms(
            schedules=schedules,
            start_date=request.start_date,
            end_date=request.end_date,
            default_user_id=request.user_id,
        )

        for schedule in schedules:
            self._validate_time_fields(
                start_date=request.start_date,
                end_date=request.end_date,
                start_time=schedule.get("start_time"),
                end_time=schedule.get("end_time"),
            )
            await self._validate_schedule_conflicts(
                user_id=int(schedule.get("user_id") or request.user_id),
                room_id=int(schedule["room_id"]),
                day_of_week=int(schedule["day_of_week"]),
                start_date=request.start_date,
                end_date=request.end_date,
                start_period=int(schedule["start_period"]),
                number_of_periods=int(schedule["number_of_periods"]),
            )

        data = request.model_dump(exclude={"schedules"})
        first_schedule = schedules[0]
        data["day_of_week"] = first_schedule["day_of_week"]
        data["start_period"] = first_schedule["start_period"]
        data["number_of_periods"] = first_schedule["number_of_periods"]
        data["start_time"] = first_schedule.get("start_time")
        data["end_time"] = first_schedule.get("end_time")
        data["room_id"] = first_schedule["room_id"]
        data["name"] = normalized_name
        data["is_cancel"] = False

        try:
            created = await self.repo.create(data)
        except IntegrityError as exc:
            if self._is_name_unique_violation(exc):
                raise AlreadyExists(
                    ERROR_CODES.COURSE_SECTION.COURSE_SECTION_NAME_IS_EXISTED
                ) from exc
            raise
        await self.repo.replace_schedules(created.id, schedules)
        await self._sync_class_sessions(
            section_id=created.id,
            start_date=request.start_date,
            end_date=request.end_date,
            schedules=schedules,
        )
        refreshed = await self.repo.get_by_id(created.id)
        return refreshed if refreshed is not None else created

    async def update(
        self,
        section_id: int,
        request: CourseSectionUpdateRequest,
    ) -> CourseSection:
        item = await self.get_by_id(section_id)
        data = request.model_dump(exclude_unset=True, exclude={"schedules"})

        if "name" in data and data["name"] is not None:
            normalized_name = data["name"].strip()
            existed = await self.repo.get_by_name_ci(normalized_name)
            if existed is not None and existed.id != item.id:
                raise AlreadyExists(
                    ERROR_CODES.COURSE_SECTION.COURSE_SECTION_NAME_IS_EXISTED
                )
            data["name"] = normalized_name

        await self._validate_foreign_keys(
            course_id=data.get("course_id"),
            user_id=data.get("user_id"),
            room_id=data.get("room_id"),
        )

        start_date = data.get("start_date", item.start_date)
        end_date = data.get("end_date", item.end_date)
        user_id = data.get("user_id", item.user_id)

        schedules = self._normalize_schedules_from_update(
            request=request,
            item=item,
            data=data,
        )
        if not schedules:
            raise Validation(message="Lớp tín chỉ phải có ít nhất một lịch học")
        schedules = self._apply_period_times(
            reference_date=start_date,
            schedules=schedules,
        )

        self._validate_internal_schedule_items(schedules)
        await self._validate_schedule_users(schedules, user_id)
        await self._validate_schedule_rooms(
            schedules=schedules,
            start_date=start_date,
            end_date=end_date,
            default_user_id=user_id,
            exclude_section_id=item.id,
        )

        for schedule in schedules:
            self._validate_time_fields(
                start_date=start_date,
                end_date=end_date,
                start_time=schedule.get("start_time"),
                end_time=schedule.get("end_time"),
            )

            await self._validate_schedule_conflicts(
                user_id=int(schedule.get("user_id") or user_id),
                room_id=int(schedule["room_id"]),
                day_of_week=int(schedule["day_of_week"]),
                start_date=start_date,
                end_date=end_date,
                start_period=int(schedule["start_period"]),
                number_of_periods=int(schedule["number_of_periods"]),
                exclude_section_id=item.id,
            )

        first_schedule = schedules[0]
        data["day_of_week"] = first_schedule["day_of_week"]
        data["start_period"] = first_schedule["start_period"]
        data["number_of_periods"] = first_schedule["number_of_periods"]
        data["start_time"] = first_schedule.get("start_time")
        data["end_time"] = first_schedule.get("end_time")
        data["room_id"] = first_schedule["room_id"]

        try:
            updated = await self.repo.update(item, data)
        except IntegrityError as exc:
            if self._is_name_unique_violation(exc):
                raise AlreadyExists(
                    ERROR_CODES.COURSE_SECTION.COURSE_SECTION_NAME_IS_EXISTED
                ) from exc
            raise
        await self.repo.replace_schedules(updated.id, schedules)
        await self._sync_class_sessions(
            section_id=updated.id,
            start_date=start_date,
            end_date=end_date,
            schedules=schedules,
        )
        refreshed = await self.repo.get_by_id(updated.id)
        return refreshed if refreshed is not None else updated

    async def delete(self, section_id: int) -> CourseSection:
        item = await self.get_by_id(section_id)
        return await self.repo.soft_delete(item)

    async def get_form_options(self):
        courses = await self.repo.list_course_options()
        lecturers = await self.repo.list_lecturer_options()
        rooms = await self.repo.list_room_options()
        return courses, lecturers, rooms

    async def list_enrolled_students(
        self,
        section_id: int,
        pagination: PaginationParams,
        search: Optional[str] = None,
    ) -> tuple[list[Student], int]:
        await self.get_by_id(section_id)

        students = await self.repo.list_enrolled_students(
            section_id=section_id,
            skip=pagination.offset,
            limit=pagination.limit,
            search=search,
        )
        total = await self.repo.count_enrolled_students(
            section_id=section_id, search=search
        )
        return students, total

    async def add_student_to_section(
        self,
        section_id: int,
        request: CourseSectionEnrollmentCreateRequest,
    ) -> Student:
        await self.get_by_id(section_id)

        student = await self.repo.get_student_by_id(request.student_id)
        if student is None:
            raise NotFound(ERROR_CODES.COURSE_SECTION.STUDENT_NOT_FOUND)

        enrollment = await self.repo.get_enrollment(
            student_id=request.student_id,
            section_id=section_id,
            include_cancel=True,
        )
        if enrollment is not None and not enrollment.is_cancel:
            raise AlreadyExists(ERROR_CODES.COURSE_SECTION.STUDENT_ALREADY_ENROLLED)

        if enrollment is not None and enrollment.is_cancel:
            await self.repo.restore_enrollment(enrollment)
        else:
            await self.repo.create_enrollment(
                student_id=request.student_id, section_id=section_id
            )

        refreshed = await self.repo.get_student_by_id(request.student_id)
        return refreshed if refreshed is not None else student

    async def remove_student_from_section(
        self, section_id: int, student_id: int
    ) -> None:
        await self.get_by_id(section_id)

        student = await self.repo.get_student_by_id(student_id)
        if student is None:
            raise NotFound(ERROR_CODES.COURSE_SECTION.STUDENT_NOT_FOUND)

        enrollment = await self.repo.get_enrollment(
            student_id=student_id,
            section_id=section_id,
            include_cancel=False,
        )
        if enrollment is None:
            raise Validation(
                message="Sinh viên chưa thuộc lớp tín chỉ này",
                error_code=ERROR_CODES.COURSE_SECTION.STUDENT_NOT_ENROLLED,
            )

        await self.repo.soft_delete_enrollment(enrollment)

    async def build_student_import_template(self, section_id: int) -> bytes:
        await self.get_by_id(section_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "credit_class_students"
        ws.append(["Mã sinh viên"])
        ws.append(["22A1001D0043"])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.read()

    async def import_students_from_excel(
        self,
        section_id: int,
        file_content: bytes,
        filename: str | None = None,
    ) -> StudentImportResultResponse:
        await self.get_by_id(section_id)

        if not file_content:
            raise Validation(message="File Excel trống")

        self._validate_file_extension(filename)

        try:
            wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise Validation(
                message="Không thể đọc file Excel. Vui lòng kiểm tra định dạng .xlsx"
            ) from exc

        ws = wb.active
        if ws.max_row < 2:
            raise Validation(message="File Excel không có dữ liệu sinh viên")

        raw_headers = [
            self._stringify_cell(cell)
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        header_index = self._build_header_index(raw_headers)

        if "student_code" not in header_index:
            raise Validation(message="Thiếu cột bắt buộc trong file mẫu: Mã sinh viên")

        errors: list[StudentImportErrorResponse] = []
        candidates: list[tuple[int, str, str]] = []
        seen_codes_in_file: set[str] = set()
        non_empty_rows = 0

        for row_idx, row_values in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            student_code_raw = self._stringify_cell(
                row_values[header_index["student_code"]]
            )

            if not any(self._stringify_cell(cell) for cell in row_values):
                continue

            non_empty_rows += 1

            if not student_code_raw:
                self._append_import_error(
                    errors,
                    row_idx,
                    "student_code",
                    "Mã sinh viên là bắt buộc",
                )
                continue

            code_key = student_code_raw.lower()
            if code_key in seen_codes_in_file:
                self._append_import_error(
                    errors,
                    row_idx,
                    "student_code",
                    "Mã sinh viên bị trùng trong file import",
                    student_code_raw,
                )
                continue

            seen_codes_in_file.add(code_key)
            candidates.append((row_idx, student_code_raw, code_key))

        imported_count = 0
        for row_idx, student_code_raw, _code_key in candidates:
            student = await self.repo.get_student_by_code_ci(student_code_raw)
            if student is None:
                self._append_import_error(
                    errors,
                    row_idx,
                    "student_code",
                    "Mã sinh viên không tồn tại trong hệ thống",
                    student_code_raw,
                )
                continue

            enrollment = await self.repo.get_enrollment(
                student_id=student.id,
                section_id=section_id,
                include_cancel=True,
            )

            if enrollment is not None and not enrollment.is_cancel:
                self._append_import_error(
                    errors,
                    row_idx,
                    "student_code",
                    "Sinh viên đã tồn tại trong lớp tín chỉ",
                    student_code_raw,
                )
                continue

            if enrollment is not None and enrollment.is_cancel:
                await self.repo.restore_enrollment(enrollment)
            else:
                await self.repo.create_enrollment(
                    student_id=student.id,
                    section_id=section_id,
                )
            imported_count += 1

        failed_count = len(errors)
        return StudentImportResultResponse(
            total_rows=non_empty_rows,
            imported_count=imported_count,
            failed_count=failed_count,
            errors=errors,
        )
