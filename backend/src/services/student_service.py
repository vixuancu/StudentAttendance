"""
StudentService – Business logic thuần cho Student.
KHÔNG import FastAPI/HTTP. Chỉ throw BusinessException.
"""

from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
import logging
import re
import unicodedata
from typing import Any, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from src.db.models.administrative_class import AdministrativeClass
from src.db.models.student import Student
from src.dto.common import PaginationParams
from src.dto.request.student_request import StudentCreateRequest, StudentUpdateRequest
from src.dto.response.student_response import (
    StudentImportErrorResponse,
    StudentImportResultResponse,
    StudentStatsResponse,
)
from src.repository.interfaces.i_student_repo import IStudentRepository
from src.services.interfaces.i_student_service import IStudentService
from src.utils.exceptions import AlreadyExistsException, NotFoundException, ValidationException


@dataclass
class _ImportRowCandidate:
    row: int
    student_code_raw: str
    student_code_key: str
    full_name: str
    birth_of_date: datetime | None
    gender: bool | None
    administrative_class_name_raw: str


class StudentService(IStudentService):
    logger = logging.getLogger(__name__)
    IMPORT_BATCH_SIZE = 500

    HEADER_ALIASES = {
        "student_code": {
            "masinhvien",
            "mssv",
            "studentcode",
            "code",
        },
        "full_name": {
            "hoten",
            "hovaten",
            "fullname",
            "tensinhvien",
        },
        "birth_of_date": {
            "ngaysinh",
            "birthday",
            "birthofdate",
            "dob",
            "dateofbirth",
        },
        "gender": {
            "gioitinh",
            "gender",
            "sex",
            "gt",
        },
        "administrative_class": {
            "lophanhchinh",
            "lop",
            "class",
            "administrativeclass",
            "classname",
        },
    }

    DATE_TEXT_FORMATS = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
        "%m/%d/%Y",
        "%d/%m/%y",
        "%d-%m-%y",
    )

    GENDER_TRUE_VALUES = {
        "1",
        "true",
        "male",
        "nam",
        "m",
    }
    GENDER_FALSE_VALUES = {
        "0",
        "false",
        "female",
        "nu",
        "f",
    }
    GENDER_NULL_VALUES = {
        "",
        "none",
        "null",
        "khac",
        "other",
        "unknown",
        "chuacapnhat",
        "chuarox",
        "-",
    }

    def __init__(self, repo: IStudentRepository):
        self.repo = repo

    @staticmethod
    def _normalize_text(value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value.strip().lower())
        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        normalized = re.sub(r"[^a-z0-9]+", "", normalized)
        return normalized

    @staticmethod
    def _stringify_cell(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _parse_gender(self, value: Any) -> bool | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            if float(value) == 1.0:
                return True
            if float(value) == 0.0:
                return False

        text = self._normalize_text(self._stringify_cell(value))
        if text in self.GENDER_NULL_VALUES:
            return None
        if text in self.GENDER_TRUE_VALUES:
            return True
        if text in self.GENDER_FALSE_VALUES:
            return False
        raise ValueError("Giới tính không hợp lệ")

    def _parse_birth_of_date(self, value: Any) -> datetime | None:
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.replace(tzinfo=None)

        if isinstance(value, date):
            return datetime.combine(value, time.min)

        if isinstance(value, (int, float)):
            dt = from_excel(value)
            if isinstance(dt, datetime):
                return dt.replace(tzinfo=None)
            if isinstance(dt, date):
                return datetime.combine(dt, time.min)

        text = self._stringify_cell(value)
        if not text:
            return None

        iso_candidate = text.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(iso_candidate)
            return dt.replace(tzinfo=None)
        except ValueError:
            pass

        for fmt in self.DATE_TEXT_FORMATS:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue

        raise ValueError("Ngày sinh không đúng định dạng")

    def _build_header_index(self, headers: list[str]) -> dict[str, int]:
        mapped: dict[str, int] = {}
        for idx, raw in enumerate(headers):
            key = self._normalize_text(raw)
            if not key:
                continue
            for column, aliases in self.HEADER_ALIASES.items():
                if key in aliases and column not in mapped:
                    mapped[column] = idx
                    break
        return mapped

    @staticmethod
    def _append_import_error(
        errors: list[StudentImportErrorResponse],
        row: int,
        field: str,
        message: str,
        student_code: str | None = None,
    ) -> None:
        errors.append(
            StudentImportErrorResponse(
                row=row,
                field=field,
                student_code=student_code,
                message=message,
            )
        )

    @staticmethod
    def _is_unique_violation(exc: Exception) -> bool:
        text = str(exc).lower()
        return "unique" in text or "duplicate" in text

    async def _insert_import_batch_with_fallback(
        self,
        rows: list[tuple[_ImportRowCandidate, dict]],
        errors: list[StudentImportErrorResponse],
    ) -> int:
        if not rows:
            return 0

        try:
            await self.repo.bulk_insert([payload for _, payload in rows])
            return len(rows)
        except Exception:  # noqa: BLE001
            self.logger.exception(
                "Bulk import fallback activated for %d rows",
                len(rows),
            )

        inserted = 0
        for item, payload in rows:
            try:
                await self.repo.create(payload)
                inserted += 1
            except Exception as exc:  # noqa: BLE001
                self.logger.warning(
                    "Import row failed row=%s code=%s err=%s",
                    item.row,
                    item.student_code_raw,
                    exc,
                )
                message = "Mã sinh viên đã tồn tại trong hệ thống"
                if not self._is_unique_violation(exc):
                    message = "Không thể lưu bản ghi do lỗi dữ liệu"
                self._append_import_error(
                    errors,
                    item.row,
                    "student_code",
                    message,
                    item.student_code_raw,
                )
        return inserted

    async def get_by_id(self, id: int) -> Student:
        student = await self.repo.get_by_id(id)
        if not student:
            raise NotFoundException(resource="Sinh viên", identifier=id)
        return student

    async def get_students(
        self,
        pagination: PaginationParams,
        search: Optional[str] = None,
        administrative_class_id: Optional[int] = None,
        is_cancel: Optional[bool] = None,
    ) -> Tuple[List[Student], int]:
        filters: List[Any] = []

        if is_cancel is not None:
            filters.append(Student.is_cancel.is_(is_cancel))

        if search:
            keyword = f"%{search.strip()}%"
            filters.append(
                Student.full_name.ilike(keyword) | Student.student_code.ilike(keyword)
            )
        if administrative_class_id:
            filters.append(Student.administrative_class_id == administrative_class_id)

        total = await self.repo.count(filters=filters)
        students = await self.repo.get_all(
            skip=pagination.offset,
            limit=pagination.limit,
            filters=filters,
        )

        student_ids = [student.id for student in students]
        face_count_map = await self.repo.count_faces_by_student_ids(student_ids)
        for student in students:
            setattr(student, "face_count", face_count_map.get(student.id, 0))

        return students, total

    async def get_student_stats(
        self,
        search: Optional[str] = None,
        administrative_class_id: Optional[int] = None,
    ) -> StudentStatsResponse:
        base_filters: List[Any] = []
        if search:
            keyword = f"%{search.strip()}%"
            base_filters.append(
                Student.full_name.ilike(keyword) | Student.student_code.ilike(keyword)
            )
        if administrative_class_id:
            base_filters.append(Student.administrative_class_id == administrative_class_id)

        total = await self.repo.count_with_filters(base_filters)
        active_count = await self.repo.count_with_filters(
            [*base_filters, Student.is_cancel.is_(False)]
        )
        locked_count = await self.repo.count_with_filters(
            [*base_filters, Student.is_cancel.is_(True)]
        )

        return StudentStatsResponse(
            total=total,
            active_count=active_count,
            locked_count=locked_count,
        )

    async def create(self, request: StudentCreateRequest) -> Student:
        student_code = request.student_code.strip()
        existing_codes = await self.repo.get_existing_student_codes_ci([student_code])
        if student_code.lower() in existing_codes:
            raise AlreadyExistsException(
                resource="Sinh viên",
                field="student_code",
                value=student_code,
            )

        admin_class = await self.repo.get_administrative_class_by_id(request.administrative_class_id)
        if admin_class is None:
            raise ValidationException("Lớp hành chính không hợp lệ", field="administrative_class_id")

        payload = request.model_dump(exclude_none=True)
        payload["student_code"] = student_code.upper()
        created = await self.repo.create(payload)
        refreshed = await self.repo.get_by_id(created.id)
        return refreshed if refreshed is not None else created

    async def update(self, id: int, request: StudentUpdateRequest) -> Student:
        student = await self.get_by_id(id)
        update_data = request.model_dump(exclude_unset=True)

        if "administrative_class_id" in update_data and update_data["administrative_class_id"] is not None:
            admin_class = await self.repo.get_administrative_class_by_id(update_data["administrative_class_id"])
            if admin_class is None:
                raise ValidationException("Lớp hành chính không hợp lệ", field="administrative_class_id")

        updated = await self.repo.update(student, update_data)
        refreshed = await self.repo.get_by_id(updated.id)
        return refreshed if refreshed is not None else updated

    async def delete(self, id: int) -> bool:
        student = await self.get_by_id(id)
        return await self.repo.update(student, {"is_cancel": True})

    async def hard_delete(self, id: int) -> bool:
        student = await self.get_by_id(id)
        return await self.repo.delete(student.id)

    async def get_administrative_class_options(self) -> List[AdministrativeClass]:
        return await self.repo.get_all_administrative_classes()

    def build_import_template(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "students"

        headers = [
            "Mã sinh viên",
            "Họ và tên",
            "Ngày sinh",
            "Giới tính",
            "Lớp hành chính",
        ]
        ws.append(headers)
        ws.append(["22A1001D0043", "Nguyễn Văn A", "2004-01-22", "Nam", "2210A01"])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.read()

    @staticmethod
    def _validate_file_extension(filename: str | None) -> None:
        if not filename:
            return
        lower_name = filename.lower()
        if lower_name.endswith(".xlsx"):
            return
        if lower_name.endswith(".xls"):
            raise ValidationException(
                "Định dạng .xls chưa được hỗ trợ, vui lòng lưu file dưới dạng .xlsx",
                field="file",
            )
        raise ValidationException(
            "Chỉ hỗ trợ file Excel định dạng .xlsx",
            field="file",
        )

    async def import_students_from_excel(
        self,
        file_content: bytes,
        filename: str | None = None,
    ) -> StudentImportResultResponse:
        if not file_content:
            raise ValidationException("File Excel trống", field="file")

        self._validate_file_extension(filename)

        try:
            wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ValidationException(
                "Không thể đọc file Excel. Vui lòng kiểm tra định dạng .xlsx",
                field="file",
            ) from exc

        ws = wb.active
        if ws.max_row < 2:
            raise ValidationException(
                "File Excel không có dữ liệu sinh viên",
                field="file",
            )

        raw_headers = [self._stringify_cell(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        header_index = self._build_header_index(raw_headers)

        required_header_keys = ("student_code", "full_name", "administrative_class")
        missing_required = [k for k in required_header_keys if k not in header_index]
        if missing_required:
            raise ValidationException(
                "Thiếu cột bắt buộc trong file mẫu: Mã sinh viên, Họ và tên, Lớp hành chính",
                field="header",
            )

        errors: list[StudentImportErrorResponse] = []
        candidates: list[_ImportRowCandidate] = []
        seen_codes_in_file: set[str] = set()
        non_empty_rows = 0

        for row_idx, row_values in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            student_code_raw = self._stringify_cell(row_values[header_index["student_code"]])
            full_name_raw = self._stringify_cell(row_values[header_index["full_name"]])
            class_name_raw = self._stringify_cell(row_values[header_index["administrative_class"]])

            birth_raw = None
            if "birth_of_date" in header_index:
                birth_raw = row_values[header_index["birth_of_date"]]

            gender_raw = None
            if "gender" in header_index:
                gender_raw = row_values[header_index["gender"]]

            if not any(
                [
                    student_code_raw,
                    full_name_raw,
                    class_name_raw,
                    self._stringify_cell(birth_raw),
                    self._stringify_cell(gender_raw),
                ]
            ):
                continue

            non_empty_rows += 1

            student_code_key = student_code_raw.lower()

            if not student_code_raw:
                self._append_import_error(errors, row_idx, "student_code", "Mã sinh viên là bắt buộc")
                continue
            if len(student_code_raw) > 100:
                self._append_import_error(
                    errors,
                    row_idx,
                    "student_code",
                    "Mã sinh viên vượt quá 100 ký tự",
                    student_code_raw,
                )
                continue
            if not full_name_raw:
                self._append_import_error(
                    errors,
                    row_idx,
                    "full_name",
                    "Họ và tên là bắt buộc",
                    student_code_raw,
                )
                continue
            if len(full_name_raw) > 100:
                self._append_import_error(
                    errors,
                    row_idx,
                    "full_name",
                    "Họ và tên vượt quá 100 ký tự",
                    student_code_raw,
                )
                continue
            if not class_name_raw:
                self._append_import_error(
                    errors,
                    row_idx,
                    "administrative_class",
                    "Lớp hành chính là bắt buộc",
                    student_code_raw,
                )
                continue

            if student_code_key in seen_codes_in_file:
                self._append_import_error(
                    errors,
                    row_idx,
                    "student_code",
                    "Mã sinh viên bị trùng trong file import",
                    student_code_raw,
                )
                continue
            seen_codes_in_file.add(student_code_key)

            try:
                birth_of_date = self._parse_birth_of_date(birth_raw)
            except ValueError as exc:
                self._append_import_error(
                    errors,
                    row_idx,
                    "birth_of_date",
                    str(exc),
                    student_code_raw,
                )
                continue

            try:
                gender = self._parse_gender(gender_raw)
            except ValueError as exc:
                self._append_import_error(
                    errors,
                    row_idx,
                    "gender",
                    str(exc),
                    student_code_raw,
                )
                continue

            candidates.append(
                _ImportRowCandidate(
                    row=row_idx,
                    student_code_raw=student_code_raw,
                    student_code_key=student_code_key,
                    full_name=full_name_raw,
                    birth_of_date=birth_of_date,
                    gender=gender,
                    administrative_class_name_raw=class_name_raw,
                )
            )

        if not candidates:
            return StudentImportResultResponse(
                total_rows=non_empty_rows,
                imported_count=0,
                failed_count=len(errors),
                errors=errors,
            )

        all_classes = await self.repo.get_all_administrative_classes()
        classes_map = {
            self._normalize_text(item.name): item
            for item in all_classes
            if item.name and item.name.strip()
        }
        existing_codes = await self.repo.get_existing_student_codes_ci(
            [item.student_code_raw for item in candidates]
        )

        imported_count = 0
        rows_to_insert: list[tuple[_ImportRowCandidate, dict]] = []
        for item in candidates:
            if item.student_code_key in existing_codes:
                self._append_import_error(
                    errors,
                    item.row,
                    "student_code",
                    "Mã sinh viên đã tồn tại trong hệ thống",
                    item.student_code_raw,
                )
                continue

            admin_class = classes_map.get(self._normalize_text(item.administrative_class_name_raw))
            if admin_class is None:
                self._append_import_error(
                    errors,
                    item.row,
                    "administrative_class",
                    "Lớp hành chính không tồn tại trong hệ thống",
                    item.student_code_raw,
                )
                continue

            payload = {
                "student_code": item.student_code_raw.upper(),
                "full_name": item.full_name,
                "birth_of_date": item.birth_of_date,
                "gender": item.gender,
                "administrative_class_id": admin_class.id,
                "is_cancel": False,
            }

            rows_to_insert.append((item, payload))

        if rows_to_insert:
            for start in range(0, len(rows_to_insert), self.IMPORT_BATCH_SIZE):
                batch = rows_to_insert[start : start + self.IMPORT_BATCH_SIZE]
                imported_count += await self._insert_import_batch_with_fallback(batch, errors)

        failed_count = len(errors)
        return StudentImportResultResponse(
            total_rows=non_empty_rows,
            imported_count=imported_count,
            failed_count=failed_count,
            errors=errors,
        )
